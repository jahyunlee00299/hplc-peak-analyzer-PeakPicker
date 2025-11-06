"""
Excel export module for peak analysis results
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import numpy as np

try:
    from .peak_detector import Peak
except ImportError:
    Peak = None


class ExcelExporter:
    """Export peak analysis results to Excel"""

    def __init__(self, output_dir: str = "results"):
        """
        Initialize Excel exporter

        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True, parents=True)

    def export_peaks(
        self,
        peaks: List,
        filename: str,
        sample_name: str,
        metadata: Optional[Dict] = None
    ) -> str:
        """
        Export peaks to Excel file

        Args:
            peaks: List of Peak objects
            filename: Base filename
            sample_name: Sample name
            metadata: Additional metadata dictionary

        Returns:
            Path to saved Excel file
        """
        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Add sheets
        self._add_metadata_sheet(wb, sample_name, peaks, metadata)
        self._add_peak_data_sheet(wb, peaks)
        self._add_summary_sheet(wb, peaks)

        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"{filename}_peaks_{timestamp}.xlsx"
        wb.save(output_file)

        return str(output_file)

    def _add_metadata_sheet(
        self,
        wb: Workbook,
        sample_name: str,
        peaks: List,
        metadata: Optional[Dict]
    ):
        """Add metadata sheet"""
        ws = wb.create_sheet("Metadata", 0)

        # Title
        ws['A1'] = "Peak Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)

        # Metadata
        row = 3
        metadata_items = [
            ("Sample Name", sample_name),
            ("Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Number of Peaks", len(peaks)),
            ("Total Area", sum(p.area for p in peaks) if peaks else 0),
        ]

        if metadata:
            metadata_items.extend(metadata.items())

        for key, value in metadata_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def _add_peak_data_sheet(self, wb: Workbook, peaks: List):
        """Add peak data sheet"""
        ws = wb.create_sheet("Peak Data")

        # Headers
        headers = [
            "Peak #",
            "RT (min)",
            "RT Start (min)",
            "RT End (min)",
            "Height",
            "Area",
            "Width (min)",
            "% Area"
        ]

        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Write peak data
        for peak_num, peak in enumerate(peaks, 1):
            row = peak_num + 1
            ws.cell(row, 1, peak_num)
            ws.cell(row, 2, round(peak.rt, 4))
            ws.cell(row, 3, round(peak.rt_start, 4))
            ws.cell(row, 4, round(peak.rt_end, 4))
            ws.cell(row, 5, round(peak.height, 2))
            ws.cell(row, 6, round(peak.area, 2))
            ws.cell(row, 7, round(peak.width, 4))
            ws.cell(row, 8, round(peak.percent_area, 2))

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(peaks) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

    def _add_summary_sheet(self, wb: Workbook, peaks: List):
        """Add summary sheet"""
        ws = wb.create_sheet("Summary")

        if not peaks:
            ws['A1'] = "No peaks detected"
            return

        # Calculate statistics
        total_area = sum(p.area for p in peaks)
        avg_width = np.mean([p.width for p in peaks])
        avg_height = np.mean([p.height for p in peaks])
        rt_range = (min(p.rt for p in peaks), max(p.rt for p in peaks))

        # Summary data
        summary_data = [
            ("Total Peaks", len(peaks)),
            ("Total Area", round(total_area, 2)),
            ("Average Peak Width (min)", round(avg_width, 4)),
            ("Average Peak Height", round(avg_height, 2)),
            ("RT Range (min)", f"{rt_range[0]:.2f} - {rt_range[1]:.2f}"),
        ]

        # Write summary
        for row, (key, value) in enumerate(summary_data, 1):
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def export_batch_results(
        self,
        batch_results: Dict[str, List],
        output_filename: str
    ) -> str:
        """
        Export batch analysis results

        Args:
            batch_results: Dictionary mapping sample names to peak lists
            output_filename: Output filename

        Returns:
            Path to saved Excel file
        """
        # Create DataFrame with all samples
        data = []

        for sample_name, peaks in batch_results.items():
            for i, peak in enumerate(peaks, 1):
                data.append({
                    'Sample': sample_name,
                    'Peak #': i,
                    'RT (min)': round(peak.rt, 4),
                    'Height': round(peak.height, 2),
                    'Area': round(peak.area, 2),
                    'Width (min)': round(peak.width, 4),
                    '% Area': round(peak.percent_area, 2)
                })

        df = pd.DataFrame(data)

        # Save to Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"{output_filename}_batch_{timestamp}.xlsx"

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Batch Results', index=False)

            # Add summary sheet
            summary_df = df.groupby('Sample').agg({
                'Peak #': 'count',
                'Area': 'sum',
                'Height': 'mean'
            }).reset_index()
            summary_df.columns = ['Sample', 'Peak Count', 'Total Area', 'Avg Height']
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        return str(output_file)

    def export_comparison(
        self,
        sample_peaks: Dict[str, List],
        target_rts: List[float],
        output_filename: str,
        rt_tolerance: float = 0.1
    ) -> str:
        """
        Export peak comparison table

        Args:
            sample_peaks: Dictionary mapping sample names to peak lists
            target_rts: List of target retention times
            output_filename: Output filename
            rt_tolerance: RT matching tolerance

        Returns:
            Path to saved Excel file
        """
        # Create comparison table
        data = []

        for sample_name, peaks in sample_peaks.items():
            row = {'Sample': sample_name}

            for target_rt in target_rts:
                # Find closest peak
                closest_peak = None
                min_distance = float('inf')

                for peak in peaks:
                    distance = abs(peak.rt - target_rt)
                    if distance < min_distance and distance <= rt_tolerance:
                        min_distance = distance
                        closest_peak = peak

                # Add to row
                col_name = f"RT_{target_rt:.2f}"
                if closest_peak:
                    row[col_name] = round(closest_peak.area, 2)
                else:
                    row[col_name] = 0

            data.append(row)

        df = pd.DataFrame(data)

        # Save to Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"{output_filename}_comparison_{timestamp}.xlsx"
        df.to_excel(output_file, index=False)

        return str(output_file)
