"""
EXPORT.CSV 파일로 피크 검출하고 엑셀 리포트 생성
Signal 시트 + Report 시트
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from peak_detector import PeakDetector
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path
import os


def parse_export_csv(filepath):
    """
    EXPORT.CSV 파일 파싱
    탭으로 구분된 형식: Time \t Intensity
    """
    data = []
    # 여러 인코딩 시도
    encodings = ['utf-16', 'utf-8', 'cp1252', 'latin1']
    f = None
    for encoding in encodings:
        try:
            f = open(filepath, 'r', encoding=encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if f is None:
        raise ValueError(f"Could not decode file {filepath} with any known encoding")

    with f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            # 탭으로 분리
            parts = line.split('\t')
            if len(parts) >= 2:
                try:
                    # 공백 제거하고 숫자 추출
                    time_str = parts[0].replace(' ', '')
                    intensity_str = parts[1].replace(' ', '')

                    time = float(time_str)
                    intensity = float(intensity_str)
                    data.append((time, intensity))
                except ValueError:
                    continue

    if not data:
        raise ValueError("No valid data found in EXPORT.CSV")

    time, intensity = zip(*data)
    return np.array(time), np.array(intensity)


def create_excel_with_signal_and_report(
    time,
    intensity,
    peaks,
    output_filename,
    sample_name="Sample"
):
    """
    엑셀 파일 생성
    - Sheet 1: Signal (원본 크로마토그램 데이터)
    - Sheet 2: Report (피크 분석 결과)
    """

    # Signal 시트용 데이터프레임
    signal_df = pd.DataFrame({
        'Time (min)': time,
        'Intensity': intensity
    })

    # Report 시트용 데이터프레임
    report_data = []
    for i, peak in enumerate(peaks, 1):
        report_data.append({
            'Peak #': i,
            'RT (min)': round(peak.rt, 3),
            'Start (min)': round(peak.rt_start, 3),
            'End (min)': round(peak.rt_end, 3),
            'Height': round(peak.height, 2),
            'Area': round(peak.area, 2),
            'Width (min)': round(peak.width, 3),
            '% Area': 0.0  # 나중에 계산
        })

    report_df = pd.DataFrame(report_data)

    # % Area 계산
    if len(peaks) > 0:
        total_area = report_df['Area'].sum()
        report_df['% Area'] = (report_df['Area'] / total_area * 100).round(2)

    # 엑셀 파일 생성
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # Sheet 1: Signal
        signal_df.to_excel(writer, sheet_name='Signal', index=False)

        # Sheet 2: Report
        # 헤더 정보
        ws_report = writer.book.create_sheet('Report')

        # 샘플 정보
        ws_report['A1'] = 'HPLC Analysis Report'
        ws_report['A1'].font = Font(size=16, bold=True)
        ws_report['A3'] = 'Sample Name:'
        ws_report['B3'] = sample_name
        ws_report['A4'] = 'Analysis Date:'
        ws_report['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_report['A5'] = 'Number of Peaks:'
        ws_report['B5'] = len(peaks)
        ws_report['A6'] = 'Total Area:'
        ws_report['B6'] = round(total_area, 2) if len(peaks) > 0 else 0

        # 피크 데이터 (7행부터 시작)
        ws_report['A8'] = 'Peak Results'
        ws_report['A8'].font = Font(size=14, bold=True)

        # 테이블 헤더 (9행)
        headers = ['Peak #', 'RT (min)', 'Start (min)', 'End (min)',
                   'Height', 'Area', 'Width (min)', '% Area']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_report.cell(row=9, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 피크 데이터 삽입 (10행부터)
        for row_idx, row_data in enumerate(report_df.values, 10):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_report.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # 교대로 색상 적용
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 열 너비 조정
        column_widths = [10, 12, 12, 12, 15, 15, 12, 12]
        for col_idx, width in enumerate(column_widths, 1):
            ws_report.column_dimensions[chr(64 + col_idx)].width = width

    print(f"[OK] Excel file created: {output_filename}")
    return output_filename


def main():
    print("=" * 70)
    print("EXPORT.CSV 파일로 피크 검출 및 엑셀 리포트 생성")
    print("=" * 70)

    # 출력 폴더 생성
    output_dir = Path("analysis_results")
    output_dir.mkdir(exist_ok=True)
    print(f"\n출력 폴더: {output_dir.absolute()}")

    # 1. EXPORT.CSV 파일 읽기
    print("\n[1] EXPORT.CSV 파일 로딩 중...")
    filepath = 'peakpicker/examples/EXPORT.CSV'
    time, intensity = parse_export_csv(filepath)

    print(f"   - 데이터 포인트: {len(time)}")
    print(f"   - 시간 범위: {time[0]:.3f} ~ {time[-1]:.3f} 분")
    print(f"   - 강도 범위: {intensity.min():.2f} ~ {intensity.max():.2f}")
    print(f"   - 평균 강도: {np.mean(intensity):.2f}")
    print(f"   - 표준편차: {np.std(intensity):.2f}")

    # 2. 피크 검출 파라미터 자동 계산
    intensity_mean = np.mean(intensity)
    intensity_std = np.std(intensity)

    # 데이터 특성에 맞게 조정
    prominence = max(2 * intensity_std, 50)  # 최소값 50
    min_height = max(intensity_mean + intensity_std, 100)  # 최소값 100

    print(f"\n[2] 피크 검출 파라미터:")
    print(f"   - prominence: {prominence:.1f}")
    print(f"   - min_height: {min_height:.1f}")
    print(f"   - min_width: 0.01 분")

    # 3. 피크 검출
    print(f"\n[3] 피크 검출 중...")
    detector = PeakDetector(
        time=time,
        intensity=intensity,
        prominence=prominence,
        min_height=min_height,
        min_width=0.01,
    )
    peaks = detector.detect_peaks()

    print(f"   - 검출된 피크 수: {len(peaks)}")

    if len(peaks) > 0:
        print(f"\n   검출된 피크 상세:")
        print(f"   {'No':<4} {'RT':<10} {'Height':<12} {'Area':<15} {'Width':<10}")
        print(f"   {'-' * 55}")
        for i, peak in enumerate(peaks, 1):
            print(f"   {i:<4} {peak.rt:<10.3f} {peak.height:<12.1f} {peak.area:<15.2f} {peak.width:<10.3f}")

        # 통계
        summary = detector.get_summary()
        print(f"\n[4] 통계:")
        print(f"   - 총 면적: {summary['total_area']:.2f}")
        print(f"   - 평균 피크 높이: {summary['avg_peak_height']:.1f}")
        print(f"   - 평균 피크 폭: {summary['avg_peak_width']:.3f} 분")
    else:
        print("   - 피크가 검출되지 않았습니다!")
        return

    # 4. 엑셀 파일 생성 (Signal + Report)
    print(f"\n[5] 엑셀 리포트 생성 중...")
    output_file = output_dir / "EXPORT_Analysis_Report.xlsx"
    create_excel_with_signal_and_report(
        time=time,
        intensity=intensity,
        peaks=peaks,
        output_filename=str(output_file),
        sample_name="EXPORT Sample"
    )

    # 5. 크로마토그램 그래프 생성
    print(f"\n[6] 크로마토그램 그래프 생성 중...")

    # 피크 영역 계산 (여유 공간 추가)
    margin_time = 0.5  # 피크 전후 0.5분 여유
    rt_min = min(p.rt_start for p in peaks) - margin_time
    rt_max = max(p.rt_end for p in peaks) + margin_time
    rt_min = max(rt_min, time[0])  # 데이터 범위 내로 제한
    rt_max = min(rt_max, time[-1])

    # 해당 영역만 추출
    mask = (time >= rt_min) & (time <= rt_max)
    time_plot = time[mask]
    intensity_plot = intensity[mask]

    fig, ax = plt.subplots(figsize=(14, 7))

    # 크로마토그램
    ax.plot(time_plot, intensity_plot, 'b-', linewidth=1.2, label='Signal', alpha=0.7)

    # 각 피크별로 다른 색상 정의
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F']

    # 피크 마커
    peak_times = np.array([p.rt for p in peaks])
    peak_intensities = np.array([intensity[p.index] for p in peaks])

    # 피크 영역 표시
    for i, peak in enumerate(peaks, 1):
        color = colors[i % len(colors)]

        # 베이스라인
        baseline_x = [time[peak.index_start], time[peak.index_end]]
        baseline_y = [intensity[peak.index_start], intensity[peak.index_end]]
        ax.plot(baseline_x, baseline_y, '--', linewidth=2, alpha=0.8, color=color)

        # 적분 영역 (각 피크별 색상)
        peak_time = time[peak.index_start:peak.index_end+1]
        peak_int = intensity[peak.index_start:peak.index_end+1]
        baseline_int = np.linspace(intensity[peak.index_start],
                                   intensity[peak.index_end],
                                   len(peak_int))
        ax.fill_between(peak_time, baseline_int, peak_int, alpha=0.4, color=color,
                       label=f'Peak {i}')

        # 피크 마커
        ax.plot(peak.rt, intensity[peak.index], 'o', markersize=10, color=color,
               markeredgecolor='white', markeredgewidth=2, zorder=5)

        # 피크 번호 + RT 표기
        label_text = f'Peak {i}\nRT: {peak.rt:.2f} min'
        ax.text(peak.rt, intensity[peak.index] + intensity_plot.max() * 0.05, label_text,
                ha='center', va='bottom', fontsize=10, fontweight='bold',
                color=color, bbox=dict(boxstyle='round,pad=0.5',
                facecolor='white', edgecolor=color, linewidth=2, alpha=0.9))

    ax.set_xlabel('Retention Time (min)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Intensity', fontsize=12, fontweight='bold')
    ax.set_title(f'HPLC Chromatogram - EXPORT.CSV ({len(peaks)} Peaks Detected)',
                fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(loc='upper right', fontsize=10, framealpha=0.9)

    plt.tight_layout()
    plot_file = output_dir / 'EXPORT_chromatogram.png'
    plt.savefig(plot_file, dpi=300, bbox_inches='tight')
    print(f"   - 그래프 저장: {plot_file}")

    print(f"\n{'=' * 70}")
    print(f"완료!")
    print(f"생성된 파일 ({output_dir}):")
    print(f"  1. EXPORT_Analysis_Report.xlsx (Signal 시트 + Report 시트)")
    print(f"  2. EXPORT_chromatogram.png (크로마토그램 그래프)")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
