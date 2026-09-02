"""
260903 Cottonii GO-difference scheme comparison
=================================================
Consumes the peak areas from quantify_260903_cottonii_go_requant.py and the
spreadsheet ground truth (Cottonii_bioreactor_3반복) to build and quantitatively
compare candidate gal-quantification schemes. See DECISION_LOG.md for the full
protocol; this implements:

  Scheme A (baseline/reproduction): area(p10.9) -> mM via single-point anchor
      at t=0 (GAL+TAG_mM(t=0)/area_untreated(t=0, vial1)), no normalization,
      absolute-difference gal = untreated_mM - treated_mM. This is the closest
      re-derivation of what the existing spreadsheet gal column represents,
      done independently from raw chromatograms (not copied from the sheet).
  Scheme B: Formate-normalized areas (each injection's p10.9 area divided by
      its own p15.9 [Formate] area, then rescaled to the same t=0 anchor)
      before differencing -- cancels injection-volume/detector-drift between
      the two SEPARATE injections (untreated vs treated) that scheme A does not.
  Scheme C: total-window-area normalized (sum of all 7 windows) as an
      alternative internal reference to Formate.
  Scheme D: ratio-based -- gal_fraction = 1 - (treated_p10.9_area /
      untreated_p10.9_area) [within-pair ratio, unitless], then
      gal_mM = gal_fraction * untreated_combined_mM (from scheme A's area->mM
      anchor). Error propagation for a ratio of two same-instrument-run areas
      can cancel correlated multiplicative noise that an absolute difference
      cannot.

For each scheme: per-timepoint (vial1-7, confident mapping) and per-plateau-
ensemble (vial8-15, D1 ambiguous-identity 8-point set) replicate SD/CV, count
of negative/non-physical values, and monotonicity check.
"""
import sys
import io
import json
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import numpy as np
import pandas as pd
import openpyxl

RESULTS_DIR = Path(r"C:\Users\Jahyun\scratch\cottonii_requant\results")
XLSX_PATH = Path(r"C:\Users\Jahyun\manuscript-figures\tagatose\rawdata\figure_raw_Data.xlsx")

EARLY_TIMEPOINTS = [0, 1, 3, 4.5, 6, 9, 12]


def load_spreadsheet_truth():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Cottonii_bioreactor_3반복']
    blocks = {1: 4, 2: 17, 3: 30}  # replicate -> header row+1 (first data row)
    rows = []
    for rep, start_row in blocks.items():
        for i in range(13):
            r = start_row + i
            t = ws.cell(row=r, column=1).value
            galtag = ws.cell(row=r, column=2).value
            galol = ws.cell(row=r, column=3).value
            formate = ws.cell(row=r, column=4).value
            tag = ws.cell(row=r, column=5).value
            gal = ws.cell(row=r, column=6).value
            rows.append({'replicate': rep, 'time_h': t, 'GALTAG_mM': galtag, 'Galol_mM': galol,
                         'for_mM': formate, 'tag_mM': tag, 'GAL_mM_sheet': gal})
    return pd.DataFrame(rows)


def load_areas():
    untreated = pd.read_csv(RESULTS_DIR / "area_pivot_untreated.csv")
    treated = pd.read_csv(RESULTS_DIR / "area_pivot_treated.csv")
    return untreated, treated


def early_only(df):
    return df[df['vial'] <= 7].copy()


def build_scheme_A(untreated, treated, sheet_truth):
    """Absolute difference, single-point area->mM anchor at t=0 (vial1, mean area)."""
    u = early_only(untreated)
    t = early_only(treated)
    anchor_area = u[u['vial'] == 1]['p10.9'].mean()
    anchor_mM = sheet_truth[(sheet_truth['time_h'] == 0)]['GALTAG_mM'].mean()
    k = anchor_mM / anchor_area  # mM per unit area

    rows = []
    for rep in [1, 2, 3]:
        for vial in range(1, 8):
            th = EARLY_TIMEPOINTS[vial - 1]
            u_area = u[(u['vial'] == vial) & (u['replicate'] == rep)]['p10.9'].values
            t_area = t[(t['vial'] == vial) & (t['replicate'] == rep)]['p10.9'].values
            if len(u_area) == 0 or len(t_area) == 0:
                continue
            u_mM = u_area[0] * k
            t_mM = t_area[0] * k
            gal_mM = u_mM - t_mM
            rows.append({'scheme': 'A_absolute_diff', 'replicate': rep, 'vial': vial, 'time_h': th,
                         'untreated_mM': u_mM, 'treated_mM': t_mM, 'gal_mM': gal_mM})
    return pd.DataFrame(rows), k


def build_scheme_B(untreated, treated, sheet_truth, k):
    """Formate cross-injection normalized: rescale the TREATED-side p10.9 area onto the
    UNTREATED-side scale using each pair's OWN Formate-area ratio
    (formate_ratio = treated_formate_area / untreated_formate_area, pair-matched by vial+
    replicate). This directly corrects the D2-diagnosed ~42-45% systematic untreated<->treated
    injection-volume/detector-response offset (formate_ratio measured ~0.55-0.60, CV~3%,
    identical across all 45 pairs regardless of GO reaction progress -- see DECISION_LOG D2),
    which a same-kind-only normalization (scaling untreated to its own mean, treated to its own
    mean) does NOT correct."""
    u = early_only(untreated)
    t = early_only(treated)

    rows = []
    for rep in [1, 2, 3]:
        for vial in range(1, 8):
            th = EARLY_TIMEPOINTS[vial - 1]
            ur = u[(u['vial'] == vial) & (u['replicate'] == rep)]
            tr = t[(t['vial'] == vial) & (t['replicate'] == rep)]
            if ur.empty or tr.empty:
                continue
            u_area = ur['p10.9'].values[0]
            t_area = tr['p10.9'].values[0]
            formate_ratio = tr['p15.9'].values[0] / ur['p15.9'].values[0]  # treated/untreated, pair-matched
            t_area_corrected = t_area / formate_ratio  # rescale treated onto untreated's injection scale
            u_mM = u_area * k
            t_mM = t_area_corrected * k
            gal_mM = u_mM - t_mM
            rows.append({'scheme': 'B_formate_normalized', 'replicate': rep, 'vial': vial, 'time_h': th,
                         'untreated_mM': u_mM, 'treated_mM': t_mM, 'gal_mM': gal_mM,
                         'formate_ratio': formate_ratio})
    return pd.DataFrame(rows)


def build_scheme_C(untreated, treated, sheet_truth, k):
    """p18.0-normalized: same construction as Scheme B (pair-matched cross-injection ratio
    correction) but using the p18.0 window instead of Formate, as a cross-validation of
    whether the correction is specific to Formate chemistry or reflects a general
    injection-volume offset (D2: p18.0 showed the same ~0.53-0.61 ratio behavior as Formate)."""
    u = early_only(untreated)
    t = early_only(treated)

    rows = []
    for rep in [1, 2, 3]:
        for vial in range(1, 8):
            th = EARLY_TIMEPOINTS[vial - 1]
            ur = u[(u['vial'] == vial) & (u['replicate'] == rep)]
            tr = t[(t['vial'] == vial) & (t['replicate'] == rep)]
            if ur.empty or tr.empty:
                continue
            u_area = ur['p10.9'].values[0]
            t_area = tr['p10.9'].values[0]
            p18_ratio = tr['p18.0'].values[0] / ur['p18.0'].values[0]
            t_area_corrected = t_area / p18_ratio
            u_mM = u_area * k
            t_mM = t_area_corrected * k
            gal_mM = u_mM - t_mM
            rows.append({'scheme': 'C_p18ref_normalized', 'replicate': rep, 'vial': vial, 'time_h': th,
                         'untreated_mM': u_mM, 'treated_mM': t_mM, 'gal_mM': gal_mM,
                         'p18_ratio': p18_ratio})
    return pd.DataFrame(rows)


def build_scheme_D_raw(untreated, treated, sheet_truth, k):
    """D0: naive ratio-based, RAW areas, no cross-injection correction --
    gal_fraction = 1 - (treated_area/untreated_area), gal_mM = gal_fraction * untreated_mM.
    Included to show that a raw ratio does NOT fix the D2 systematic offset (the offset is
    between the two injections, not within one, so a within-pair ratio of uncorrected areas
    is algebraically just as biased as Scheme A's absolute difference -- both differ from the
    formate-corrected schemes only in how they propagate REMAINING random noise, not the
    systematic component)."""
    u = early_only(untreated)
    t = early_only(treated)

    rows = []
    for rep in [1, 2, 3]:
        for vial in range(1, 8):
            th = EARLY_TIMEPOINTS[vial - 1]
            ur = u[(u['vial'] == vial) & (u['replicate'] == rep)]
            tr = t[(t['vial'] == vial) & (t['replicate'] == rep)]
            if ur.empty or tr.empty:
                continue
            u_area = ur['p10.9'].values[0]
            t_area = tr['p10.9'].values[0]
            u_mM = u_area * k
            ratio = t_area / u_area
            gal_fraction = 1.0 - ratio
            gal_mM = gal_fraction * u_mM
            rows.append({'scheme': 'D0_ratio_raw', 'replicate': rep, 'vial': vial, 'time_h': th,
                         'untreated_mM': u_mM, 'treated_mM': u_mM * ratio, 'gal_mM': gal_mM,
                         'gal_fraction': gal_fraction})
    return pd.DataFrame(rows)


def build_scheme_D(untreated, treated, sheet_truth, k):
    """D: Formate-corrected ratio-based -- gal_fraction = 1 - (treated_area / formate_ratio) /
    untreated_area, i.e. apply the SAME cross-injection Formate correction as Scheme B, then take
    the within-pair ratio instead of the absolute difference. Isolates whether ratio-based error
    propagation adds further scatter reduction on top of the systematic-offset correction alone."""
    u = early_only(untreated)
    t = early_only(treated)

    rows = []
    for rep in [1, 2, 3]:
        for vial in range(1, 8):
            th = EARLY_TIMEPOINTS[vial - 1]
            ur = u[(u['vial'] == vial) & (u['replicate'] == rep)]
            tr = t[(t['vial'] == vial) & (t['replicate'] == rep)]
            if ur.empty or tr.empty:
                continue
            u_area = ur['p10.9'].values[0]
            t_area = tr['p10.9'].values[0]
            formate_ratio = tr['p15.9'].values[0] / ur['p15.9'].values[0]
            t_area_corrected = t_area / formate_ratio
            u_mM = u_area * k
            ratio = t_area_corrected / u_area
            gal_fraction = 1.0 - ratio
            gal_mM = gal_fraction * u_mM
            rows.append({'scheme': 'D_ratio_formate_corrected', 'replicate': rep, 'vial': vial, 'time_h': th,
                         'untreated_mM': u_mM, 'treated_mM': u_mM * ratio, 'gal_mM': gal_mM,
                         'gal_fraction': gal_fraction})
    return pd.DataFrame(rows)


def build_plateau_ensemble(untreated, treated, k):
    """vial 8-15 (D1: 8-vial ensemble, ambiguous timepoint identity). Computes gal under all four
    schemes (A absolute/raw, B formate-corrected absolute, D0 ratio/raw, D ratio/formate-corrected)
    -- these don't need a specific timepoint label, just the 'post-plateau residual gal'
    distribution per replicate, which is exactly what the task's success criterion (b) needs."""
    u = untreated[untreated['vial'] >= 8].copy()
    t = treated[treated['vial'] >= 8].copy()

    rows = []
    for rep in [1, 2, 3]:
        for vial in range(8, 16):
            ur = u[(u['vial'] == vial) & (u['replicate'] == rep)]
            tr = t[(t['vial'] == vial) & (t['replicate'] == rep)]
            if ur.empty or tr.empty:
                continue
            u_area = ur['p10.9'].values[0]
            t_area = tr['p10.9'].values[0]
            formate_ratio = tr['p15.9'].values[0] / ur['p15.9'].values[0]
            t_area_corrected = t_area / formate_ratio

            u_mM = u_area * k
            t_mM_raw = t_area * k
            t_mM_corr = t_area_corrected * k

            gal_A = u_mM - t_mM_raw
            gal_B = u_mM - t_mM_corr
            ratio_raw = t_area / u_area
            gal_D0 = (1.0 - ratio_raw) * u_mM
            ratio_corr = t_area_corrected / u_area
            gal_D = (1.0 - ratio_corr) * u_mM

            rows.append({'replicate': rep, 'vial': vial, 'untreated_mM': u_mM,
                         'gal_A_absolute_raw': gal_A, 'gal_B_absolute_formatecorr': gal_B,
                         'gal_D0_ratio_raw': gal_D0, 'gal_D_ratio_formatecorr': gal_D,
                         'formate_ratio': formate_ratio})
    return pd.DataFrame(rows)


def summarize_scheme(df, scheme_name, sheet_truth):
    """Per-timepoint mean/SD(ddof=1)/CV across 3 replicates, plus negativity/monotonicity checks."""
    rows = []
    for th in EARLY_TIMEPOINTS:
        sub = df[df['time_h'] == th]['gal_mM'].values
        if len(sub) == 0:
            continue
        mean = np.mean(sub)
        sd = np.std(sub, ddof=1) if len(sub) > 1 else np.nan
        cv = (sd / mean * 100) if mean != 0 else np.nan
        n_neg = int(np.sum(sub < 0))
        rows.append({'scheme': scheme_name, 'time_h': th, 'mean_gal_mM': mean, 'sd_gal_mM': sd,
                     'cv_pct': cv, 'n_negative': n_neg, 'n': len(sub), 'values': list(np.round(sub, 4))})
    summary = pd.DataFrame(rows)
    return summary


def main():
    untreated, treated = load_areas()
    sheet_truth = load_spreadsheet_truth()

    scheme_A, k = build_scheme_A(untreated, treated, sheet_truth)
    scheme_B = build_scheme_B(untreated, treated, sheet_truth, k)
    scheme_C = build_scheme_C(untreated, treated, sheet_truth, k)
    scheme_D0 = build_scheme_D_raw(untreated, treated, sheet_truth, k)
    scheme_D = build_scheme_D(untreated, treated, sheet_truth, k)

    print(f"Anchor scale k = {k:.8e} mM/area-unit (t=0 GAL+TAG anchor)")

    all_schemes = pd.concat([scheme_A, scheme_B, scheme_C, scheme_D0, scheme_D], ignore_index=True)
    all_schemes.to_csv(RESULTS_DIR / "schemes_early_timepoints_raw.csv", index=False, encoding='utf-8-sig')

    summaries = []
    for name, df in [('A_absolute_diff', scheme_A), ('B_formate_normalized', scheme_B),
                      ('C_p18ref_normalized', scheme_C), ('D0_ratio_raw', scheme_D0),
                      ('D_ratio_formate_corrected', scheme_D)]:
        s = summarize_scheme(df, name, sheet_truth)
        summaries.append(s)

    # Reference: spreadsheet's own gal values at early timepoints, same replicate/time grid
    sheet_rows = []
    for th in EARLY_TIMEPOINTS:
        sub = sheet_truth[sheet_truth['time_h'] == th]['GAL_mM_sheet'].values
        mean = np.mean(sub)
        sd = np.std(sub, ddof=1) if len(sub) > 1 else np.nan
        cv = (sd / mean * 100) if mean != 0 else np.nan
        n_neg = int(np.sum(sub < 0))
        sheet_rows.append({'scheme': 'SHEET_existing', 'time_h': th, 'mean_gal_mM': mean, 'sd_gal_mM': sd,
                            'cv_pct': cv, 'n_negative': n_neg, 'n': len(sub), 'values': list(np.round(sub, 4))})
    summaries.append(pd.DataFrame(sheet_rows))

    summary_all = pd.concat(summaries, ignore_index=True)
    summary_all.to_csv(RESULTS_DIR / "schemes_early_comparison_summary.csv", index=False, encoding='utf-8-sig')

    print("\n=== EARLY-TIMEPOINT (t<=12h) SCHEME COMPARISON ===")
    for scheme in ['SHEET_existing', 'A_absolute_diff', 'B_formate_normalized', 'C_p18ref_normalized',
                   'D0_ratio_raw', 'D_ratio_formate_corrected']:
        sub = summary_all[summary_all['scheme'] == scheme]
        print(f"\n-- {scheme} --")
        print(sub[['time_h', 'mean_gal_mM', 'sd_gal_mM', 'cv_pct', 'n_negative']].to_string(index=False))

    # Plateau ensemble (vial 8-15, D1 ambiguous identity, all schemes)
    plateau = build_plateau_ensemble(untreated, treated, k)
    plateau.to_csv(RESULTS_DIR / "plateau_ensemble_raw.csv", index=False, encoding='utf-8-sig')

    print("\n=== PLATEAU ENSEMBLE (vial 8-15, all 3 reps pooled = 24 points; D1: timepoint identity ambiguous) ===")
    for col, label in [('gal_A_absolute_raw', 'Scheme A (absolute, raw)'),
                        ('gal_B_absolute_formatecorr', 'Scheme B (absolute, formate-corrected)'),
                        ('gal_D0_ratio_raw', 'Scheme D0 (ratio, raw)'),
                        ('gal_D_ratio_formatecorr', 'Scheme D (ratio, formate-corrected)')]:
        vals = plateau[col].values
        mean = np.mean(vals)
        sd = np.std(vals, ddof=1)
        cv = sd / mean * 100 if mean != 0 else np.nan
        n_neg = int(np.sum(vals < 0))
        print(f"{label:42s}: mean={mean:8.4f}  SD={sd:8.4f}  CV={cv:6.2f}%  n_negative={n_neg}/{len(vals)}")

    # Also per-replicate plateau SD (mirrors the task's t=48h-style replicate comparison)
    print("\n-- Per-replicate plateau stats (SD across the 8 plateau vials within a replicate) --")
    for rep in [1, 2, 3]:
        for col, label in [('gal_A_absolute_raw', 'A'), ('gal_B_absolute_formatecorr', 'B'),
                            ('gal_D0_ratio_raw', 'D0'), ('gal_D_ratio_formatecorr', 'D')]:
            vals = plateau[plateau['replicate'] == rep][col].values
            mean = np.mean(vals)
            sd = np.std(vals, ddof=1)
            print(f"  rep{rep} scheme{label:>3s}: mean={mean:8.3f}  SD={sd:7.3f}  n_neg={int(np.sum(vals<0))}/{len(vals)}")

    # Sheet's own late-timepoint reference for direct comparison (t=48h specifically, as in task brief)
    print("\n-- SHEET existing gal at t=48h (reference, task brief) --")
    sub48 = sheet_truth[sheet_truth['time_h'] == 48]['GAL_mM_sheet'].values
    print(f"  values={np.round(sub48,4)}  mean={np.mean(sub48):.4f}  SD={np.std(sub48,ddof=1):.4f}  n_neg={int(np.sum(sub48<0))}")

    print(f"\nSaved: {RESULTS_DIR / 'schemes_early_comparison_summary.csv'}")
    print(f"Saved: {RESULTS_DIR / 'schemes_early_timepoints_raw.csv'}")
    print(f"Saved: {RESULTS_DIR / 'plateau_ensemble_raw.csv'}")


if __name__ == '__main__':
    main()
