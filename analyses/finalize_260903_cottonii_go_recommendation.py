"""
260903 Cottonii GO-difference: final comparison table + diagnostic figures
=============================================================================
Consumes outputs from quantify_260903_cottonii_go_requant.py,
compare_260903_cottonii_go_schemes.py, and errorprop_260903_cottonii_go.py to
produce:
  1. A single comparison table: sheet vs scheme A (raw) vs scheme B (formate-
     corrected), all timepoints (early confident + plateau ensemble), with
     SD/CV/negativity/monotonicity flags -- the artifact the final
     recommendation is based on.
  2. Diagnostic figures: overlaid untreated/treated chromatograms at early/
     mid/late vials, and per-scheme gal time-course with error bars.
"""
import sys
import io
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from rainbow.agilent.chemstation import parse_ch
import re

DATA_DIR = Path(r"C:\Users\Jahyun\scratch\cot7z\260624_cottonii_bioreactor")
RESULTS_DIR = Path(r"C:\Users\Jahyun\scratch\cottonii_requant\results")
FIG_DIR = Path(r"C:\Users\Jahyun\scratch\cottonii_requant\figures")
FIG_DIR.mkdir(parents=True, exist_ok=True)

EARLY_TIMEPOINTS = [0, 1, 3, 4.5, 6, 9, 12]


def vial_key(name):
    m = re.match(r'(?:260625_)?COTTONII_REACTOR(?:_GO)?_(\d+)_(\d+)\.D', name)
    return int(m.group(1)), int(m.group(2))


def read_rid(ch_path):
    r = parse_ch(str(ch_path))
    time = np.asarray(r.xlabels, dtype=float)
    sig = np.asarray(r.data, dtype=float).flatten()
    n = min(len(time), len(sig))
    return time[:n], sig[:n]


def build_final_table():
    early = pd.read_csv(RESULTS_DIR / "schemes_early_comparison_summary.csv")
    plateau = pd.read_csv(RESULTS_DIR / "plateau_ensemble_raw.csv")

    rows = []
    for th in EARLY_TIMEPOINTS:
        sheet = early[(early.scheme == 'SHEET_existing') & (early.time_h == th)]
        a = early[(early.scheme == 'A_absolute_diff') & (early.time_h == th)]
        b = early[(early.scheme == 'B_formate_normalized') & (early.time_h == th)]
        rows.append({
            'segment': f't={th}h', 'time_h': th,
            'sheet_mean': sheet.mean_gal_mM.values[0], 'sheet_sd': sheet.sd_gal_mM.values[0],
            'sheet_cv_pct': sheet.cv_pct.values[0], 'sheet_n_neg': int(sheet.n_negative.values[0]),
            'schemeA_mean': a.mean_gal_mM.values[0], 'schemeA_sd': a.sd_gal_mM.values[0],
            'schemeA_cv_pct': a.cv_pct.values[0], 'schemeA_n_neg': int(a.n_negative.values[0]),
            'schemeB_mean': b.mean_gal_mM.values[0], 'schemeB_sd': b.sd_gal_mM.values[0],
            'schemeB_cv_pct': b.cv_pct.values[0], 'schemeB_n_neg': int(b.n_negative.values[0]),
        })

    # plateau ensemble (8-15, pooled 24 pts) as one aggregate "row", plus per-replicate breakdown
    a_vals = plateau['gal_A_absolute_raw'].values
    b_vals = plateau['gal_B_absolute_formatecorr'].values
    rows.append({
        'segment': 'plateau(vial8-15,n=24,t>=15h ambiguous-identity ensemble; D1)', 'time_h': np.nan,
        'sheet_mean': np.nan, 'sheet_sd': np.nan, 'sheet_cv_pct': np.nan, 'sheet_n_neg': np.nan,
        'schemeA_mean': np.mean(a_vals), 'schemeA_sd': np.std(a_vals, ddof=1),
        'schemeA_cv_pct': np.std(a_vals, ddof=1) / np.mean(a_vals) * 100, 'schemeA_n_neg': int(np.sum(a_vals < 0)),
        'schemeB_mean': np.mean(b_vals), 'schemeB_sd': np.std(b_vals, ddof=1),
        'schemeB_cv_pct': np.std(b_vals, ddof=1) / np.mean(b_vals) * 100, 'schemeB_n_neg': int(np.sum(b_vals < 0)),
    })

    # sheet's own late timepoints for direct reference (not vial-mapped, just its own values)
    import openpyxl
    wb = openpyxl.load_workbook(r"C:\Users\Jahyun\manuscript-figures\tagatose\rawdata\figure_raw_Data.xlsx", data_only=True)
    ws = wb['Cottonii_bioreactor_3반복']
    blocks = {1: 4, 2: 17, 3: 30}
    late_ts = [15, 18, 24, 36, 48, 66]
    for th in late_ts:
        vals = []
        for rep, start_row in blocks.items():
            for i in range(13):
                r = start_row + i
                t = ws.cell(row=r, column=1).value
                if t == th:
                    vals.append(ws.cell(row=r, column=6).value)
        vals = np.array(vals)
        rows.append({
            'segment': f'SHEET_only t={th}h', 'time_h': th,
            'sheet_mean': np.mean(vals), 'sheet_sd': np.std(vals, ddof=1) if len(vals) > 1 else np.nan,
            'sheet_cv_pct': np.std(vals, ddof=1) / np.mean(vals) * 100 if np.mean(vals) != 0 else np.nan,
            'sheet_n_neg': int(np.sum(vals < 0)),
            'schemeA_mean': np.nan, 'schemeA_sd': np.nan, 'schemeA_cv_pct': np.nan, 'schemeA_n_neg': np.nan,
            'schemeB_mean': np.nan, 'schemeB_sd': np.nan, 'schemeB_cv_pct': np.nan, 'schemeB_n_neg': np.nan,
        })

    df = pd.DataFrame(rows)
    out_path = RESULTS_DIR / "FINAL_comparison_table.csv"
    df.to_csv(out_path, index=False, encoding='utf-8-sig')
    print(f"Saved final comparison table: {out_path}")
    print(df.round(3).to_string(index=False))
    return df


def plot_chromatogram_overlays():
    """Overlaid untreated vs GO-treated chromatograms at early/mid/late vials (rep1)."""
    vials = [1, 4, 8, 15]
    fig, axes = plt.subplots(len(vials), 1, figsize=(10, 3.2 * len(vials)), sharex=True)

    for ax, v in zip(axes, vials):
        u_folder = DATA_DIR / f"COTTONII_REACTOR_{v}_1.D"
        t_folder = DATA_DIR / f"260625_COTTONII_REACTOR_GO_{v}_1.D"
        tu, yu = read_rid(u_folder / "RID1A.ch")
        tt, yt = read_rid(t_folder / "RID1A.ch")

        mask_u = (tu >= 9.5) & (tu <= 17.0)
        mask_t = (tt >= 9.5) & (tt <= 17.0)
        ax.plot(tu[mask_u], yu[mask_u], color='#3498db', linewidth=1.0, label='untreated (GAL+TAG)')
        ax.plot(tt[mask_t], yt[mask_t], color='#e74c3c', linewidth=1.0, label='GO-treated (TAG only)')
        ax.axvspan(10.5, 11.5, color='gray', alpha=0.08)
        ax.axvspan(15.5, 16.3, color='gold', alpha=0.12)
        ax.set_ylabel('nRIU', fontsize=9)
        ax.set_title(f"vial {v} rep1 (untreated={u_folder.name}, treated={t_folder.name})", fontsize=8, loc='left')
        ax.legend(fontsize=7, loc='upper right')
        ax.grid(alpha=0.2, linewidth=0.5)

    axes[-1].set_xlabel('Retention Time (min)', fontsize=10)
    fig.suptitle('Cottonii GO-difference: untreated vs GO-treated overlay (early/mid/late vials)\n'
                 'gray=p10.9 combined/tag window, gold=p15.9 Formate reference window', fontsize=11, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    out_path = FIG_DIR / "chromatogram_overlays_early_mid_late.png"
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved: {out_path}")


def plot_gal_timecourse():
    """Per-scheme gal time-course with error bars, vs sheet's own values."""
    early = pd.read_csv(RESULTS_DIR / "schemes_early_comparison_summary.csv")
    plateau = pd.read_csv(RESULTS_DIR / "plateau_ensemble_raw.csv")

    fig, ax = plt.subplots(figsize=(10, 6))

    sheet_early = early[early.scheme == 'SHEET_existing']
    a_early = early[early.scheme == 'A_absolute_diff']
    b_early = early[early.scheme == 'B_formate_normalized']

    ax.errorbar(sheet_early.time_h, sheet_early.mean_gal_mM, yerr=sheet_early.sd_gal_mM,
                marker='o', color='black', label='Sheet (existing)', capsize=3, linewidth=1.3)
    ax.errorbar(a_early.time_h, a_early.mean_gal_mM, yerr=a_early.sd_gal_mM,
                marker='s', color='#e74c3c', label='Scheme A (raw absolute diff, this reanalysis)',
                capsize=3, linewidth=1.3, alpha=0.85)
    ax.errorbar(b_early.time_h, b_early.mean_gal_mM, yerr=b_early.sd_gal_mM,
                marker='^', color='#27ae60', label='Scheme B (formate-corrected, this reanalysis)',
                capsize=3, linewidth=1.3, alpha=0.85)

    # plateau ensemble plotted at a nominal x=40h (midpoint of the ambiguous 15-66h range) as a
    # single aggregate point, clearly annotated as such
    plateau_x = 40
    a_mean, a_sd = plateau.gal_A_absolute_raw.mean(), plateau.gal_A_absolute_raw.std(ddof=1)
    b_mean, b_sd = plateau.gal_B_absolute_formatecorr.mean(), plateau.gal_B_absolute_formatecorr.std(ddof=1)
    ax.errorbar([plateau_x], [a_mean], yerr=[a_sd], marker='s', color='#e74c3c', capsize=4,
                markersize=10, markerfacecolor='none', markeredgewidth=2)
    ax.errorbar([plateau_x], [b_mean], yerr=[b_sd], marker='^', color='#27ae60', capsize=4,
                markersize=10, markerfacecolor='none', markeredgewidth=2)
    ax.annotate('plateau ensemble\n(vial8-15, n=24,\ntimepoint identity\nambiguous -- D1)',
                xy=(plateau_x, max(a_mean, b_mean) + max(a_sd, b_sd) + 3), fontsize=7, ha='center')

    # sheet's own late points for reference
    import openpyxl
    wb = openpyxl.load_workbook(r"C:\Users\Jahyun\manuscript-figures\tagatose\rawdata\figure_raw_Data.xlsx", data_only=True)
    ws = wb['Cottonii_bioreactor_3반복']
    blocks = {1: 4, 2: 17, 3: 30}
    late_ts = [15, 18, 24, 36, 48, 66]
    late_mean, late_sd = [], []
    for th in late_ts:
        vals = []
        for rep, start_row in blocks.items():
            for i in range(13):
                r = start_row + i
                t = ws.cell(row=r, column=1).value
                if t == th:
                    vals.append(ws.cell(row=r, column=6).value)
        vals = np.array(vals)
        late_mean.append(np.mean(vals))
        late_sd.append(np.std(vals, ddof=1))
    ax.errorbar(late_ts, late_mean, yerr=late_sd, marker='o', color='black', linestyle=':',
                capsize=3, linewidth=1.0, alpha=0.6)

    ax.axhline(0, color='gray', linewidth=0.8, linestyle='--')
    ax.set_xlabel('Time (h)', fontsize=11)
    ax.set_ylabel('Residual D-galactose (mM)', fontsize=11)
    ax.set_title('Cottonii GO-difference gal time-course: sheet vs raw vs formate-corrected reanalysis', fontsize=12)
    ax.legend(fontsize=9, loc='upper right')
    ax.grid(alpha=0.25)
    fig.tight_layout()
    out_path = FIG_DIR / "gal_timecourse_scheme_comparison.png"
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved: {out_path}")


def plot_formate_ratio_diagnostic():
    """The D2 finding: formate/p18.0 ratio is flat ~0.55-0.60 across all vials, confirming a
    systematic (not random) cross-injection offset."""
    untreated = pd.read_csv(RESULTS_DIR / "area_pivot_untreated.csv")
    treated = pd.read_csv(RESULTS_DIR / "area_pivot_treated.csv")
    merged = untreated.merge(treated, on=['vial', 'replicate'], suffixes=('_u', '_t'))
    merged['formate_ratio'] = merged['p15.9_t'] / merged['p15.9_u']
    merged['p18_ratio'] = merged['p18.0_t'] / merged['p18.0_u']

    fig, ax = plt.subplots(figsize=(9, 5))
    for rep, color in zip([1, 2, 3], ['#3498db', '#e67e22', '#9b59b6']):
        sub = merged[merged.replicate == rep].sort_values('vial')
        ax.plot(sub.vial, sub.formate_ratio, marker='o', color=color, label=f'rep{rep} Formate ratio', linewidth=1.2)
        ax.plot(sub.vial, sub.p18_ratio, marker='x', color=color, linestyle='--', alpha=0.6, label=f'rep{rep} p18.0 ratio')

    ax.axhline(1.0, color='gray', linestyle=':', linewidth=1.0, label='ratio=1 (no offset)')
    ax.set_xlabel('Vial', fontsize=11)
    ax.set_ylabel('treated / untreated area ratio', fontsize=11)
    ax.set_title('D2: systematic cross-injection offset (chemically-inert reference peaks)\n'
                 'Formate & p18.0 ratio ~0.55-0.60 across ALL vials -- not GO-reaction-dependent', fontsize=11)
    ax.legend(fontsize=7, ncol=2)
    ax.grid(alpha=0.25)
    fig.tight_layout()
    out_path = FIG_DIR / "formate_offset_diagnostic.png"
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved: {out_path}")


def main():
    build_final_table()
    plot_chromatogram_overlays()
    plot_gal_timecourse()
    plot_formate_ratio_diagnostic()


if __name__ == '__main__':
    main()
